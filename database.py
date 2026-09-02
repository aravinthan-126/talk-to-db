"""
Database helpers for Talk-to-DB.
Supports:
  1. SQLite (file upload or local path)
  2. MySQL / MariaDB
  3. PostgreSQL
  4. MongoDB + GridFS (files)
"""

import re
import sqlite3
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple, List, Any
from urllib.parse import quote_plus

import pandas as pd
from sqlalchemy import create_engine, text, inspect
from sqlalchemy.engine import Engine

SAMPLE_DB_PATH = Path(__file__).parent / "employees.db"

# ---------------------------------------------------------------------------
# Mongo / GridFS optional import
# ---------------------------------------------------------------------------
try:
    from pymongo import MongoClient
    from pymongo.errors import ConnectionFailure
    from gridfs import GridFS
    from gridfs.errors import NoFile
    from bson import ObjectId
    MONGO_AVAILABLE = True
except ImportError:
    MONGO_AVAILABLE = False
    MongoClient = None
    GridFS = None
    ObjectId = None


def _safe_str(value, default: str = "") -> str:
    if value is None:
        return default
    s = str(value).strip()
    return s if s else default


# ---------------------------------------------------------------------------
# Sample SQLite
# ---------------------------------------------------------------------------

def create_sample_database() -> str:
    if SAMPLE_DB_PATH.exists():
        return str(SAMPLE_DB_PATH)

    conn = sqlite3.connect(SAMPLE_DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE departments (
        dept_id INTEGER PRIMARY KEY,
        dept_name TEXT NOT NULL UNIQUE,
        location TEXT
    )
    """)
    cursor.execute("""
    CREATE TABLE employees (
        emp_id INTEGER PRIMARY KEY,
        first_name TEXT NOT NULL,
        last_name TEXT NOT NULL,
        email TEXT UNIQUE,
        department TEXT,
        salary REAL,
        hire_date TEXT,
        manager_id INTEGER,
        city TEXT,
        FOREIGN KEY (manager_id) REFERENCES employees(emp_id)
    )
    """)

    departments = [
        (1, "HR", "Chennai"),
        (2, "Engineering", "Bangalore"),
        (3, "Sales", "Mumbai"),
        (4, "Finance", "Chennai"),
        (5, "Marketing", "Delhi"),
        (6, "Operations", "Hyderabad"),
    ]
    cursor.executemany("INSERT INTO departments VALUES (?, ?, ?)", departments)

    employees = [
        (1, "Arun", "Kumar", "arun.kumar@company.com", "HR", 55000, "2020-03-15", None, "Chennai"),
        (2, "Priya", "Sharma", "priya.sharma@company.com", "Engineering", 85000, "2019-07-22", None, "Bangalore"),
        (3, "Rajesh", "Patel", "rajesh.patel@company.com", "Sales", 62000, "2021-01-10", None, "Mumbai"),
        (4, "Lakshmi", "Iyer", "lakshmi.iyer@company.com", "Finance", 72000, "2018-11-05", None, "Chennai"),
        (5, "Vikram", "Singh", "vikram.singh@company.com", "Engineering", 95000, "2017-05-18", 2, "Bangalore"),
        (6, "Meena", "Ravi", "meena.ravi@company.com", "HR", 48000, "2022-02-28", 1, "Chennai"),
        (7, "Suresh", "Nair", "suresh.nair@company.com", "Marketing", 58000, "2020-09-12", None, "Delhi"),
        (8, "Anitha", "Krishnan", "anitha.krishnan@company.com", "Engineering", 78000, "2021-06-30", 2, "Bangalore"),
        (9, "Karthik", "Reddy", "karthik.reddy@company.com", "Sales", 67000, "2019-12-01", 3, "Hyderabad"),
        (10, "Divya", "Menon", "divya.menon@company.com", "Finance", 69000, "2020-04-20", 4, "Chennai"),
        (11, "Mohammed", "Ali", "mohammed.ali@company.com", "Operations", 52000, "2021-08-15", None, "Hyderabad"),
        (12, "Sowmya", "Bhat", "sowmya.bhat@company.com", "Engineering", 88000, "2018-03-25", 2, "Bangalore"),
        (13, "Ganesh", "Pillai", "ganesh.pillai@company.com", "HR", 51000, "2022-07-01", 1, "Chennai"),
        (14, "Fatima", "Begum", "fatima.begum@company.com", "Marketing", 61000, "2019-10-08", 7, "Delhi"),
        (15, "Ravi", "Shankar", "ravi.shankar@company.com", "Sales", 74000, "2017-12-14", 3, "Mumbai"),
        (16, "Kavitha", "Srinivasan", "kavitha.s@company.com", "Finance", 81000, "2016-08-30", 4, "Chennai"),
        (17, "Arjun", "Das", "arjun.das@company.com", "Engineering", 92000, "2020-01-19", 2, "Bangalore"),
        (18, "Nisha", "Joshi", "nisha.joshi@company.com", "Operations", 47000, "2023-01-05", 11, "Hyderabad"),
        (19, "Sathish", "Kumar", "sathish.kumar@company.com", "Sales", 59000, "2021-11-22", 3, "Chennai"),
        (20, "Deepa", "Rao", "deepa.rao@company.com", "Marketing", 64000, "2019-04-17", 7, "Bangalore"),
    ]
    cursor.executemany(
        "INSERT INTO employees VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", employees
    )
    conn.commit()
    conn.close()
    return str(SAMPLE_DB_PATH)


# ---------------------------------------------------------------------------
# SQL engines
# ---------------------------------------------------------------------------

def create_sqlite_engine(db_path: str) -> Engine:
    return create_engine(f"sqlite:///{db_path}", future=True)


def create_mysql_engine(host: str, port: int, user: str, password: str, database: str) -> Engine:
    pwd = quote_plus(password or "")
    url = f"mysql+pymysql://{user}:{pwd}@{host}:{port}/{database}?charset=utf8mb4"
    return create_engine(url, future=True, pool_pre_ping=True)


def create_postgres_engine(host: str, port: int, user: str, password: str, database: str) -> Engine:
    pwd = quote_plus(password or "")
    url = f"postgresql+psycopg2://{user}:{pwd}@{host}:{port}/{database}"
    return create_engine(url, future=True, pool_pre_ping=True)


def test_connection(engine: Engine) -> Tuple[bool, str]:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True, "Connection successful"
    except Exception as e:
        return False, str(e)


def get_schema_from_engine(engine: Engine) -> str:
    try:
        inspector = inspect(engine)
        tables = inspector.get_table_names()
        if not tables:
            return "No tables found in the database."
        parts = ["Database Schema:\n"]
        for table in tables:
            columns = inspector.get_columns(table)
            pk = inspector.get_pk_constraint(table)
            pk_cols = set(pk.get("constrained_columns") or [])
            parts.append(f"TABLE: {table}")
            for col in columns:
                name = col["name"]
                col_type = str(col["type"])
                flags = []
                if name in pk_cols:
                    flags.append("Primary Key")
                if not col.get("nullable", True):
                    flags.append("NOT NULL")
                flag_str = f" ({', '.join(flags)})" if flags else ""
                parts.append(f"  - {name} ({col_type}){flag_str}")
            parts.append("")
        return "\n".join(parts)
    except Exception as e:
        return f"Error reading schema: {e}"


def list_tables_from_engine(engine: Engine) -> List[str]:
    try:
        return inspect(engine).get_table_names()
    except Exception:
        return []


def execute_sql(engine: Engine, sql: str) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    try:
        with engine.connect() as conn:
            df = pd.read_sql_query(text(sql), conn)
        return df, None
    except Exception as e:
        return None, str(e)


def execute_write_sql(engine: Engine, sql: str) -> Tuple[bool, str, int]:
    try:
        with engine.begin() as conn:
            result = conn.execute(text(sql))
            affected = result.rowcount if result.rowcount is not None else -1
        return True, f"✅ Query executed successfully. Rows affected: {affected}", affected
    except Exception as e:
        return False, f"❌ Error: {e}", 0


def save_uploaded_db(uploaded_file) -> str:
    suffix = Path(uploaded_file.name).suffix or ".db"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(uploaded_file.getvalue())
    tmp.close()
    return tmp.name


# ---------------------------------------------------------------------------
# MongoDB
# ---------------------------------------------------------------------------

def create_mongo_client(
    uri: str = None,
    host: str = "localhost",
    port: int = 27017,
    username: str = None,
    password: str = None,
    auth_source: str = "admin",
):
    if not MONGO_AVAILABLE:
        raise RuntimeError("pymongo is not installed. Run: pip install pymongo")
    if uri:
        return MongoClient(uri, serverSelectionTimeoutMS=8000)
    if username and password:
        return MongoClient(
            host=host,
            port=port,
            username=username,
            password=password,
            authSource=auth_source,
            serverSelectionTimeoutMS=8000,
        )
    return MongoClient(host=host, port=port, serverSelectionTimeoutMS=8000)


def test_mongo_connection(client) -> Tuple[bool, str]:
    try:
        client.admin.command("ping")
        return True, "Connection successful"
    except Exception as e:
        return False, str(e)


def list_mongo_collections(db) -> List[str]:
    try:
        return [c for c in db.list_collection_names() if not c.startswith("system.")]
    except Exception:
        return []


def get_mongo_schema(db, sample_limit: int = 2) -> str:
    try:
        collections = list_mongo_collections(db)
        if not collections:
            return "No collections found in this database."
        parts = ["MongoDB Schema (collections + sample fields):\n"]
        for name in collections:
            parts.append(f"COLLECTION: {name}")
            try:
                docs = list(db[name].find().limit(sample_limit))
                if not docs:
                    parts.append("  (empty)")
                else:
                    keys = set()
                    for d in docs:
                        keys.update(d.keys())
                    parts.append(f"  Fields: {', '.join(sorted(str(k) for k in keys))}")
                    for d in docs:
                        preview = {k: d[k] for k in list(d.keys())[:8]}
                        parts.append(f"  Sample: {preview}")
            except Exception as e:
                parts.append(f"  (error reading: {e})")
            parts.append("")
        return "\n".join(parts)
    except Exception as e:
        return f"Error reading Mongo schema: {e}"


def execute_mongo_query(db, code: str) -> Tuple[Any, Optional[str]]:
    """Execute LLM-generated pymongo code. `db` is injected."""
    if not code or not str(code).strip():
        return None, "Empty MongoDB code"
    local_vars = {"db": db, "result": None, "ObjectId": ObjectId, "datetime": datetime}
    try:
        exec(code, {"__builtins__": __builtins__}, local_vars)
        return local_vars.get("result"), None
    except Exception as e:
        return None, str(e)


def mongo_result_to_dataframe(result) -> Optional[pd.DataFrame]:
    if result is None:
        return None
    try:
        if isinstance(result, list):
            if not result:
                return pd.DataFrame()
            if all(isinstance(x, dict) for x in result):
                return pd.DataFrame(result)
            return pd.DataFrame({"value": result})
        if isinstance(result, dict):
            return pd.DataFrame([result])
        return None
    except Exception:
        return None


def format_mongo_write_result(result) -> str:
    if result is None:
        return "Done."
    if hasattr(result, "inserted_id"):
        return f"Inserted document id: {result.inserted_id}"
    if hasattr(result, "inserted_ids"):
        return f"Inserted {len(result.inserted_ids)} document(s)"
    if hasattr(result, "modified_count"):
        return f"Modified: {result.modified_count}, matched: {getattr(result, 'matched_count', '?')}"
    if hasattr(result, "deleted_count"):
        return f"Deleted: {result.deleted_count}"
    return str(result)


# ---------------------------------------------------------------------------
# GridFS  (None-safe — fixes 'NoneType'.startswith)
# ---------------------------------------------------------------------------

def upload_file_to_gridfs(
    db,
    file_bytes: bytes,
    filename: str,
    content_type: str = None,
    metadata: dict = None,
) -> str:
    fs = GridFS(db)
    file_id = fs.put(
        file_bytes,
        filename=_safe_str(filename, "unnamed"),
        content_type=_safe_str(content_type, "application/octet-stream"),
        metadata=metadata or {},
    )
    return str(file_id)


def list_gridfs_files(db, limit: int = 100) -> List[dict]:
    """All string fields are never None."""
    fs = GridFS(db)
    out = []
    for f in fs.find().sort("uploadDate", -1).limit(limit):
        ctype = getattr(f, "content_type", None)
        out.append({
            "_id": str(f._id),
            "filename": _safe_str(getattr(f, "filename", None), "unnamed"),
            "length": getattr(f, "length", 0) or 0,
            "contentType": _safe_str(ctype, "application/octet-stream"),
            "uploadDate": getattr(f, "upload_date", None),
        })
    return out


def download_from_gridfs(db, file_id: str) -> Tuple[Optional[bytes], Optional[str], Optional[str]]:
    try:
        fs = GridFS(db)
        grid_out = fs.get(ObjectId(str(file_id)))
        return (
            grid_out.read(),
            _safe_str(getattr(grid_out, "filename", None), "unnamed"),
            _safe_str(getattr(grid_out, "content_type", None), "application/octet-stream"),
        )
    except Exception:
        return None, None, None


def delete_from_gridfs(db, file_id: str) -> Tuple[bool, str]:
    try:
        fs = GridFS(db)
        fs.delete(ObjectId(str(file_id)))
        return True, "Deleted"
    except Exception as e:
        return False, str(e)


# ---------------------------------------------------------------------------
# Excel (.xlsx) support — sheets as tables, formulas via openpyxl
# ---------------------------------------------------------------------------

def load_excel_workbook(path: str) -> dict:
    """
    Load Excel file.
    Returns dict:
      path, sheets (list of names), dataframes {sheet: df}, workbook (openpyxl)
    """
    from openpyxl import load_workbook

    path = str(path)
    wb = load_workbook(path)
    sheets = wb.sheetnames
    dataframes = {}
    for name in sheets:
        try:
            dataframes[name] = pd.read_excel(path, sheet_name=name, engine="openpyxl")
        except Exception:
            dataframes[name] = pd.DataFrame()
    return {
        "path": path,
        "sheets": sheets,
        "dataframes": dataframes,
        "workbook": wb,
    }


def save_uploaded_excel(uploaded_file) -> str:
    """
    Save uploaded Excel to a persistent folder so edits survive.
    Returns absolute path. Browser cannot overwrite the user's original
    local file; user must download the updated copy.
    """
    name = Path(uploaded_file.name or "data.xlsx").name
    suffix = Path(name).suffix.lower() or ".xlsx"
    if suffix not in (".xlsx", ".xls", ".csv", ".xlsm"):
        suffix = ".xlsx"
        name = Path(name).stem + ".xlsx"

    save_dir = Path(__file__).parent / "excel_uploads"
    save_dir.mkdir(parents=True, exist_ok=True)
    # unique name to avoid clashes
    dest = save_dir / name
    if dest.exists():
        stem, suf = dest.stem, dest.suffix
        i = 1
        while True:
            cand = save_dir / f"{stem}_{i}{suf}"
            if not cand.exists():
                dest = cand
                break
            i += 1
    dest.write_bytes(uploaded_file.getvalue())
    return str(dest.resolve())


def persist_excel_dataframes(path: str, dfs: dict) -> None:
    """Write all DataFrames to the xlsx path (full replace)."""
    path = str(path)
    # Always write .xlsx
    if path.lower().endswith(".csv"):
        path = path.rsplit(".", 1)[0] + ".xlsx"
    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
        if not dfs:
            pd.DataFrame().to_excel(writer, sheet_name="Sheet1", index=False)
        else:
            for name, d in dfs.items():
                sheet = str(name)[:31] if name else "Sheet1"
                if not isinstance(d, pd.DataFrame):
                    d = pd.DataFrame(d)
                d.to_excel(writer, sheet_name=sheet, index=False)


def get_excel_schema(excel_state: dict) -> str:
    """Describe sheets + columns + sample rows for LLM."""
    parts = ["Excel Workbook Schema:\n"]
    parts.append(f"File: {excel_state.get('path', '')}\n")
    for name in excel_state.get("sheets", []):
        df = excel_state["dataframes"].get(name)
        parts.append(f"SHEET: {name}")
        if df is None or df.empty:
            parts.append("  (empty)")
        else:
            cols = [str(c) for c in df.columns]
            parts.append(f"  Columns: {', '.join(cols)}")
            parts.append(f"  Rows: {len(df)}")
            try:
                sample = df.head(3).to_dict(orient="records")
                parts.append(f"  Sample: {sample}")
            except Exception:
                pass
        parts.append("")
    parts.append(
        "You can use pandas on dataframes dict: dfs['SheetName']\n"
        "For formulas use openpyxl workbook: wb['SheetName']['A1'] = '=SUM(B2:B10)'\n"
        "After formula changes call wb.save(path).\n"
    )
    return "\n".join(parts)


def list_excel_sheets(excel_state: dict) -> List[str]:
    return list(excel_state.get("sheets") or [])



def execute_excel_code(excel_state: dict, code: str) -> Tuple[Any, Optional[str], dict]:
    """
    Execute LLM Python for Excel CRUD + formulas.
    Always saves changes to excel_state['path'] on disk.
    """
    from openpyxl import load_workbook

    if not code or not str(code).strip():
        return None, "Empty Excel code", excel_state

    path = str(excel_state["path"])
    try:
        wb = load_workbook(path)
    except Exception as e:
        return None, f"Cannot open workbook: {e}", excel_state

    dfs = {}
    for name in wb.sheetnames:
        try:
            dfs[name] = pd.read_excel(path, sheet_name=name, engine="openpyxl")
        except Exception:
            dfs[name] = pd.DataFrame()

    code_stripped = code.strip()
    if "result" not in code_stripped and code_stripped:
        lines = [ln for ln in code_stripped.splitlines() if ln.strip()]
        if lines and not lines[-1].strip().startswith("result"):
            lines[-1] = "result = " + lines[-1]
            code_stripped = "\n".join(lines)

    local_vars = {
        "dfs": dfs,
        "wb": wb,
        "path": path,
        "pd": pd,
        "result": None,
        "datetime": datetime,
        "np": __import__("numpy"),
    }
    try:
        exec(code_stripped, {"__builtins__": __builtins__}, local_vars)
        result = local_vars.get("result")
        if isinstance(local_vars.get("dfs"), dict):
            dfs = local_vars["dfs"]

        lower = code_stripped.lower()
        formula_or_cell = any(k in lower for k in ("ws[", "wb[", "=sum(", "=average(", "=if(", "=vlookup("))
        df_mutate = any(k in lower for k in (
            "to_excel", "excelwriter", ".loc[", ".iloc[", ".drop(", ".insert(",
            "concat(", ".at[", ".iat[",
        ))

        # 1) Save openpyxl formulas/cell edits
        try:
            wb.save(path)
        except Exception:
            pass

        # 2) If DataFrames were mutated, rewrite full workbook from dfs
        #    (this is what makes CRUD persist on disk)
        if df_mutate or "to_excel" in lower or "excelwriter" in lower:
            try:
                persist_excel_dataframes(path, dfs)
            except Exception as e:
                return None, f"Save failed: {e}", excel_state
        elif formula_or_cell:
            # formulas already saved via wb.save
            pass
        else:
            # pure read — no extra save needed
            pass

        # Reload state from disk so UI matches file
        try:
            new_wb = load_workbook(path)
            new_dfs = {}
            for name in new_wb.sheetnames:
                try:
                    new_dfs[name] = pd.read_excel(path, sheet_name=name, engine="openpyxl")
                except Exception:
                    new_dfs[name] = dfs.get(name, pd.DataFrame())
        except Exception:
            new_wb = wb
            new_dfs = dfs

        if result is None and new_dfs:
            result = next(iter(new_dfs.values()))

        new_state = {
            "path": path,
            "sheets": list(new_dfs.keys()),
            "dataframes": new_dfs,
            "workbook": new_wb,
            "original_name": excel_state.get("original_name") or Path(path).name,
        }
        return result, None, new_state
    except Exception as e:
        return None, str(e), excel_state



def excel_result_to_dataframe(result) -> Optional[pd.DataFrame]:
    """Convert any result into a DataFrame for table display."""
    if result is None:
        return None
    try:
        if isinstance(result, pd.DataFrame):
            return result
        if isinstance(result, pd.Series):
            return result.to_frame().reset_index()
        if isinstance(result, list):
            if not result:
                return pd.DataFrame()
            if all(isinstance(x, dict) for x in result):
                return pd.DataFrame(result)
            return pd.DataFrame({"value": result})
        if isinstance(result, dict):
            # single row or scalar map
            try:
                return pd.DataFrame([result])
            except Exception:
                return pd.DataFrame({"key": list(result.keys()), "value": list(result.values())})
        if isinstance(result, (int, float, str, bool)):
            return pd.DataFrame({"result": [result]})
        # numpy scalar
        try:
            import numpy as np
            if isinstance(result, (np.integer, np.floating)):
                return pd.DataFrame({"result": [result.item()]})
        except Exception:
            pass
        return pd.DataFrame({"result": [str(result)]})
    except Exception:
        return pd.DataFrame({"result": [str(result)]})


def is_excel_write(code: str) -> bool:
    """True when the code mutates the Excel file (data or formulas)."""
    import re
    lower = (code or "").lower()
    # Explicit write / formula markers
    hard_write = (
        "to_excel", "excelwriter", "wb.save", ".save(",
        ".drop(", ".insert(", "concat(",
        "=sum(", "=average(", "=if(", "=vlookup(", "=countif(",
        "=min(", "=max(", "=round(",
    )
    if any(k in lower for k in hard_write):
        return True
    # Cell write via openpyxl: ws['A1'] = ...
    if re.search(r"\bws\s*\[", lower) and "=" in lower:
        return True
    if re.search(r"\bwb\s*\[", lower) and "=" in lower:
        return True
    # DataFrame cell assignment: dfs['s'].loc[...] = value
    if re.search(r"dfs\s*\[[^\]]+\]\s*\.loc\s*\[[^\]]+\]\s*=", lower):
        return True
    if re.search(r"dfs\s*\[[^\]]+\]\s*\.iloc\s*\[[^\]]+\]\s*=", lower):
        return True
    if re.search(r"dfs\s*\[[^\]]+\]\s*\[[^\]]+\]\s*=", lower):
        return True
    return False

